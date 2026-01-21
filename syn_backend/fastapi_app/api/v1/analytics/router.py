import csv
import io
from pathlib import Path
from typing import Optional
from fastapi import APIRouter, Query, HTTPException, Body
from pydantic import BaseModel, Field
from fastapi.responses import StreamingResponse
from fastapi_app.core.config import settings
from myUtils.analytics_db import (
    ensure_analytics_schema,
    get_analytics_summary,
    get_analytics_videos,
    get_chart_data,
    insert_video_analytics,
    update_video_analytics,
    upsert_video_analytics_by_key,
)
from fastapi_app.api.v1.analytics.douyin_sec_uid_resolver import resolve_douyin_sec_uid as resolve_douyin_sec_uid_raw
from datetime import datetime
from typing import Any, Dict, List
import sys
import os
import httpx
import re
import asyncio
import json
import subprocess
from myUtils.cookie_manager import cookie_manager
from platforms.path_utils import resolve_cookie_file

router = APIRouter(prefix="/analytics", tags=["数据分析"])

DB_PATH: Path = Path(settings.BASE_DIR) / "db" / "database.db"
DEFAULT_UA = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)
_MEDIACRAWLER_PROFILE_IDS = {
    "kuaishou": "mediacrawler_kuaishou",
    "xiaohongshu": "mediacrawler_xiaohongshu",
}

# Ensure DB schema exists on import
ensure_analytics_schema(DB_PATH)


class CollectTask(BaseModel):
    platform: str = Field(..., description="平台：douyin/xiaohongshu/kuaishou")
    work_id: str = Field(..., description="作品 ID（aweme_id / note_id / photoId）")


class CollectPayload(BaseModel):
    mode: Optional[str] = Field(None, description="采集模式 accounts/works，默认按账号采集")
    platform: Optional[str] = Field(None, description="平台，可选：douyin/xiaohongshu/kuaishou/channels/all")
    account_ids: Optional[List[str]] = Field(None, description="账号ID列表，空为所有有效账号")
    work_ids: Optional[List[str]] = Field(None, description="作品 ID 列表，若提供 platform 则全部使用该平台")
    tasks: Optional[List[CollectTask]] = Field(None, description="自定义任务列表，含平台与作品 ID")



def _ensure_douyin_api_on_path() -> None:
    douyin_root = Path(settings.BASE_DIR) / "douyin_tiktok_api"
    if douyin_root.exists():
        sys.path.insert(0, str(douyin_root))



def _ensure_mediacrawler_cookie_file(platform: str) -> str:
    cookie_dir = Path(settings.COOKIE_FILES_DIR)
    cookie_dir.mkdir(parents=True, exist_ok=True)
    cookie_path = cookie_dir / f"mediacrawler_{platform}.json"
    if not cookie_path.exists():
        cookie_path.write_text('{"cookies":[],"origins":[]}', encoding="utf-8")
    return str(cookie_path)

def _extract_cookie_list(cookie_data: Any) -> List[Dict[str, Any]]:
    if isinstance(cookie_data, dict):
        if isinstance(cookie_data.get("cookies"), list):
            return cookie_data.get("cookies") or []
        if isinstance(cookie_data.get("cookie_info"), dict) and isinstance(cookie_data["cookie_info"].get("cookies"), list):
            return cookie_data["cookie_info"]["cookies"] or []
        if isinstance(cookie_data.get("origins"), list):
            cookies: List[Dict[str, Any]] = []
            for origin in cookie_data.get("origins") or []:
                cookies.extend(origin.get("cookies") or [])
            return cookies
    if isinstance(cookie_data, list):
        return [c for c in cookie_data if isinstance(c, dict)]
    return []


def _extract_xhs_user_id(cookie_data: Any) -> Optional[str]:
    if not isinstance(cookie_data, dict):
        return None
    origins = cookie_data.get("origins") or []
    for origin in origins:
        if not isinstance(origin, dict):
            continue
        local_storage = origin.get("localStorage") or []
        for item in local_storage:
            if not isinstance(item, dict):
                continue
            name = item.get("name")
            value = item.get("value")
            if name == "snsWebPublishCurrentUser" and value:
                return str(value)
            if name == "USER_INFO_FOR_BIZ" and value:
                try:
                    parsed = json.loads(value)
                    user_id = parsed.get("userId")
                    if user_id:
                        return str(user_id)
                except Exception:
                    continue
            if name == "USER_INFO" and value:
                try:
                    parsed = json.loads(value)
                    user_id = (parsed.get("user") or {}).get("value", {}).get("userId")
                    if user_id:
                        return str(user_id)
                except Exception:
                    continue
    return None




def _normalize_xhs_note(account_id: str, note: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    if not isinstance(note, dict):
        return None
    note_id = note.get("note_id") or note.get("id")
    if not note_id:
        return None
    cover = None
    cover_obj = note.get("cover") or {}
    if isinstance(cover_obj, dict):
        cover = (cover_obj.get("url_list") or [None])[0] or cover_obj.get("url")
    image_list = note.get("image_list") or note.get("images_list") or []
    if not cover and isinstance(image_list, list) and image_list:
        first = image_list[0] or {}
        if isinstance(first, dict):
            cover = first.get("url") or first.get("image_url")
    interact = note.get("interact_info") or note.get("interact") or {}
    return {
        "account_id": account_id,
        "platform": "xiaohongshu",
        "video_id": str(note_id),
        "video_url": f"https://www.xiaohongshu.com/explore/{note_id}",
        "title": note.get("title") or note.get("desc") or "",
        "thumbnail": cover,
        "publish_date": _to_date_str(note.get("time") or note.get("create_time")),
        "play_count": interact.get("played_count") or interact.get("expose_count") or 0,
        "like_count": interact.get("liked_count") or 0,
        "comment_count": interact.get("comment_count") or 0,
        "collect_count": interact.get("collected_count") or 0,
        "share_count": interact.get("shared_count") or 0,
        "raw_data": note,
    }


def _normalize_kuaishou_item(account_id: str, item: Dict[str, Any]) -> Optional[Dict[str, Any]]:
    if not isinstance(item, dict):
        return None
    photo = item.get("photo") if isinstance(item.get("photo"), dict) else item
    photo_id = photo.get("id") or item.get("photo_id")
    if not photo_id:
        return None
    cover = photo.get("coverUrl")
    if not cover:
        cover_urls = photo.get("coverUrls") or photo.get("cover_urls") or []
        if isinstance(cover_urls, list) and cover_urls:
            cover = cover_urls[0]
    return {
        "account_id": account_id,
        "platform": "kuaishou",
        "video_id": str(photo_id),
        "video_url": f"https://www.kuaishou.com/short-video/{photo_id}",
        "title": photo.get("caption") or photo.get("title") or "",
        "thumbnail": cover,
        "publish_date": _to_date_str(photo.get("timestamp") or photo.get("time")),
        "play_count": photo.get("viewCount") or 0,
        "like_count": photo.get("likeCount") or photo.get("realLikeCount") or 0,
        "comment_count": photo.get("commentCount") or 0,
        "collect_count": photo.get("collectCount") or 0,
        "share_count": photo.get("shareCount") or 0,
        "raw_data": item,
    }


def _build_cookie_header(cookie_data: Any, domain_filter: Optional[str] = None) -> str:
    cookies = _extract_cookie_list(cookie_data)
    pairs: List[str] = []
    for cookie in cookies:
        name = cookie.get("name")
        value = cookie.get("value")
        if not name or value is None:
            continue
        if domain_filter:
            domain = cookie.get("domain") or ""
            if domain_filter not in domain:
                continue
        pairs.append(f"{name}={value}")
    return "; ".join(pairs)


def _set_bilibili_cookie(cookie_header: str) -> None:
    if not cookie_header:
        return
    try:
        from crawlers.bilibili.web import web_crawler as bili_mod  # type: ignore
        if isinstance(getattr(bili_mod, "config", None), dict):
            bili_mod.config["TokenManager"]["bilibili"]["headers"]["cookie"] = cookie_header
    except Exception:
        return


async def _set_douyin_cookie(crawler: Any, cookie_header: str) -> None:
    if not cookie_header or crawler is None:
        return
    try:
        await crawler.update_cookie(cookie_header)
    except Exception:
        return


def _read_cookie_file_for_account(account: Dict[str, Any]) -> Dict[str, Any]:
    return cookie_manager._read_cookie_file(account.get("cookie_file") or "")


def _extract_sec_uid_from_cookie(cookie_data: Any) -> Optional[str]:
    if not isinstance(cookie_data, dict):
        return None
    user_info = cookie_data.get("user_info") or {}
    for key in ("sec_uid", "sec_user_id", "secUserId", "secUid"):
        value = user_info.get(key)
        if value:
            return str(value)
    return None


async def _resolve_douyin_sec_uid(account: Dict[str, Any], cookie_data: Dict[str, Any]) -> Optional[str]:
    """
    解析抖音账号的 sec_uid
    使用新的三层降级策略解析器
    """
    candidate = account.get("user_id") or account.get("account_id")
    if not candidate:
        return None

    cookie_header = _build_cookie_header(cookie_data, domain_filter="douyin.com") or _build_cookie_header(cookie_data)
    
    # 使用新的解析器（默认不使用 Playwright，性能优先）
    return await resolve_douyin_sec_uid_raw(
        user_id=str(candidate),
        cookie_header=cookie_header,
        use_playwright=False  # 可以通过环境变量或配置控制
    )


def _persist_douyin_sec_uid(account: Dict[str, Any], sec_uid: str) -> None:
    if not sec_uid:
        return
    try:
        cookie_file = account.get("cookie_file") or ""
        cookie_data = cookie_manager._read_cookie_file(cookie_file)
        if isinstance(cookie_data, dict):
            if "user_info" not in cookie_data:
                cookie_data["user_info"] = {}
            cookie_data["user_info"]["sec_uid"] = sec_uid
            cookie_manager._write_cookie_file(cookie_file, cookie_data)
    except Exception:
        return


def _looks_like_sec_uid(value: Optional[str]) -> bool:
    if not value:
        return False
    value = str(value)
    return value.startswith("MS4w") and len(value) > 20




def _extract_bilibili_uid(cookie_data: Any) -> Optional[str]:
    if not isinstance(cookie_data, dict):
        return None
    cookies = cookie_data.get("cookies")
    if not cookies and isinstance(cookie_data.get("cookie_info"), dict):
        cookies = cookie_data.get("cookie_info", {}).get("cookies")
    if not isinstance(cookies, list):
        return None
    for cookie in cookies:
        if isinstance(cookie, dict) and cookie.get("name") == "DedeUserID":
            value = cookie.get("value")
            if value:
                return str(value)
    return None

def _parse_bilibili_vlist(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    data = payload.get("data") if isinstance(payload, dict) else None
    if isinstance(data, dict) and "list" in data:
        data = data.get("list")
    if isinstance(data, dict):
        vlist = data.get("vlist")
        if isinstance(vlist, list):
            return vlist
    return []

def _account_lookup() -> Dict[str, Dict[str, Any]]:
    lookup: Dict[str, Dict[str, Any]] = {}
    try:
        for acc in cookie_manager.list_flat_accounts():
            account_id = str(acc.get("account_id") or "")
            if account_id:
                lookup[account_id] = acc
    except Exception:
        return {}
    return lookup


def _to_date_str(ts: Any) -> str:
    try:
        ts_int = int(ts)
        if ts_int > 1e12:
            ts_int = ts_int / 1000
        return datetime.utcfromtimestamp(ts_int).date().isoformat()
    except Exception:
        return datetime.utcnow().date().isoformat()


async def _collect_bilibili_accounts(account_ids: Optional[List[str]] = None) -> Dict[str, Any]:
    _ensure_douyin_api_on_path()
    try:
        from crawlers.bilibili.web.web_crawler import BilibiliWebCrawler  # type: ignore
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Bilibili crawler import failed: {exc}")

    accounts = [
        acc for acc in cookie_manager.list_flat_accounts()
        if acc.get("platform") == "bilibili"
    ]
    if account_ids:
        accounts = [acc for acc in accounts if acc.get("account_id") in set(account_ids)]

    crawler = BilibiliWebCrawler()
    success = 0
    failed: List[Dict[str, Any]] = []

    for account in accounts:
        account_id = account.get("account_id")
        cookie_data = cookie_manager._read_cookie_file(account.get("cookie_file") or "")
        cookie_header = _build_cookie_header(cookie_data, domain_filter="bilibili.com") or _build_cookie_header(cookie_data)
        if cookie_header:
            _set_bilibili_cookie(cookie_header)
        uid = (
            account.get("user_id")
            or cookie_manager._extract_user_id_from_cookie("bilibili", cookie_data)
            or _extract_bilibili_uid(cookie_data)
        )

        if not uid:
            failed.append({"account_id": account_id, "error": "missing uid"})
            continue

        try:
            pn = 1
            max_pages = 5
            while pn <= max_pages:
                payload = await crawler.fetch_user_post_videos(uid=str(uid), pn=pn)
                vlist = _parse_bilibili_vlist(payload)
                if not vlist:
                    break
                for item in vlist:
                    video_id = item.get("bvid") or item.get("aid")
                    if not video_id:
                        continue
                    bvid = item.get("bvid")
                    aid = item.get("aid")
                    video_url = None
                    if bvid:
                        video_url = f"https://www.bilibili.com/video/{bvid}"
                    elif aid:
                        video_url = f"https://www.bilibili.com/video/av{aid}"

                    record = {
                        "account_id": account_id,
                        "platform": "bilibili",
                        "video_id": str(video_id),
                        "video_url": video_url,
                        "title": item.get("title") or "",
                        "thumbnail": item.get("pic"),
                        "publish_date": _to_date_str(item.get("created")),
                        "play_count": item.get("play") or 0,
                        "like_count": item.get("like") or 0,
                        "comment_count": item.get("comment") or 0,
                        "collect_count": item.get("favorites") or 0,
                        "share_count": item.get("share") or 0,
                        "raw_data": item,
                    }
                    upsert_video_analytics_by_key(DB_PATH, platform="bilibili", video_id=str(video_id), data=record)
                pn += 1
            success += 1
        except Exception as exc:  # noqa: BLE001
            failed.append({"account_id": account_id, "error": str(exc)})

    return {"success": success, "failed": len(failed), "errors": failed}


def _parse_douyin_aweme_list(payload: Dict[str, Any]) -> List[Dict[str, Any]]:
    if not isinstance(payload, dict):
        return []
    data = payload.get("data") if isinstance(payload.get("data"), dict) else payload
    aweme_list = data.get("aweme_list") or []
    return aweme_list if isinstance(aweme_list, list) else []


async def _collect_douyin_tiktok_accounts(
    account_ids: Optional[List[str]] = None,
    sec_user_ids: Optional[Dict[str, str]] = None,
    max_pages: int = 5,
    page_size: int = 20,
) -> Dict[str, Any]:
    _ensure_douyin_api_on_path()
    try:
        from crawlers.douyin.web.web_crawler import DouyinWebCrawler  # type: ignore
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Douyin crawler import failed: {exc}")

    accounts = [
        acc for acc in cookie_manager.list_flat_accounts()
        if acc.get("platform") == "douyin"
    ]
    if account_ids:
        accounts = [acc for acc in accounts if acc.get("account_id") in set(account_ids)]

    success = 0
    failed: List[Dict[str, Any]] = []

    max_pages = 100

    for account in accounts:
        account_id = account.get("account_id")
        cookie_data = cookie_manager._read_cookie_file(account.get("cookie_file") or "")
        cookie_header = _build_cookie_header(cookie_data, domain_filter="douyin.com") or _build_cookie_header(cookie_data)
        crawler = DouyinWebCrawler()
        if cookie_header:
            await _set_douyin_cookie(crawler, cookie_header)

        sec_user_id = (sec_user_ids or {}).get(account_id)
        if not sec_user_id:
            sec_user_id = _extract_sec_uid_from_cookie(cookie_data)
        if not sec_user_id and _looks_like_sec_uid(account.get("user_id")):
            sec_user_id = str(account.get("user_id"))
        if not sec_user_id:
            sec_user_id = await _resolve_douyin_sec_uid(account, cookie_data)
            if sec_user_id:
                _persist_douyin_sec_uid(account, sec_user_id)
        if not sec_user_id:
            failed.append({"account_id": account_id, "error": "missing sec_user_id"})
            continue

        cursor = 0
        pages = 0
        try:
            while pages < max_pages:
                pages += 1
                payload = await crawler.fetch_user_post_videos(sec_user_id=str(sec_user_id), max_cursor=cursor, count=page_size)
                aweme_list = _parse_douyin_aweme_list(payload)
                if not aweme_list:
                    break

                for item in aweme_list:
                    if not isinstance(item, dict):
                        continue
                    stats = item.get("statistics") or {}
                    video_id = item.get("aweme_id") or ""
                    if not video_id:
                        continue
                    cover = None
                    try:
                        cover = item.get("video", {}).get("cover", {}).get("url_list", [None])[0]
                    except Exception:
                        cover = None

                    record = {
                        "account_id": account_id,
                        "platform": "douyin",
                        "video_id": str(video_id),
                        "video_url": f"https://www.douyin.com/video/{video_id}",
                        "title": item.get("desc") or item.get("title") or "",
                        "thumbnail": cover,
                        "publish_date": _to_date_str(item.get("create_time")),
                        "play_count": stats.get("play_count") or stats.get("play_count_v2") or 0,
                        "like_count": stats.get("digg_count") or 0,
                        "comment_count": stats.get("comment_count") or 0,
                        "share_count": stats.get("share_count") or 0,
                        "collect_count": stats.get("collect_count") or 0,
                        "raw_data": item,
                    }
                    upsert_video_analytics_by_key(DB_PATH, platform="douyin", video_id=str(video_id), data=record)

                data = payload.get("data") if isinstance(payload, dict) else {}
                has_more = data.get("has_more") if isinstance(data, dict) else None
                cursor = data.get("max_cursor", cursor + page_size) if isinstance(data, dict) else cursor + page_size
                if not has_more:
                    break
            success += 1
        except Exception as exc:  # noqa: BLE001
            failed.append({"account_id": account_id, "error": str(exc)})

    return {"success": success, "failed": len(failed), "errors": failed}


async def _collect_xhs_mediacrawler_accounts(account_ids: Optional[List[str]] = None) -> Dict[str, Any]:
    return {"success": 0, "failed": 0, "errors": []}


async def _collect_kuaishou_mediacrawler_accounts(account_ids: Optional[List[str]] = None) -> Dict[str, Any]:
    return {"success": 0, "failed": 0, "errors": []}

@router.get("/", summary="获取分析数据", include_in_schema=True)
@router.get("", include_in_schema=False)
async def get_analytics(
    startDate: Optional[str] = Query(None, description="Start Date YYYY-MM-DD"),
    endDate: Optional[str] = Query(None, description="End Date YYYY-MM-DD"),
    platform: Optional[str] = Query(None, description="Legacy single platform"),
    platforms: Optional[List[str]] = Query(None, description="List of platforms"),
    accounts: Optional[List[str]] = Query(None, description="List of account IDs"),
    limit: int = Query(100, ge=1, le=10000, description="Limit")
):
    """获取汇总、视频列表和图表数据"""
    try:
        # Support legacy single platform param by adding it to list if present
        if platform and platform != "all":
            if not platforms:
                platforms = []
            platforms.append(platform)

        # Build account_ids list from accounts param
        account_ids = accounts

        summary = get_analytics_summary(DB_PATH, startDate, endDate, platforms=platforms, account_ids=account_ids)
        videos = get_analytics_videos(DB_PATH, startDate, endDate, limit, platforms=platforms, account_ids=account_ids)
        
        account_map = _account_lookup()
        for video in videos:
            account_id = str(video.get("accountId") or video.get("account_id") or "")
            account = account_map.get(account_id)
            if account:
                video["accountName"] = account.get("name") or ""
                video["accountAvatar"] = account.get("avatar") or ""
                video["accountId"] = account_id
                
        # Pass filters to chart data as well
        chart_data = get_chart_data(DB_PATH, startDate, endDate, platforms=platforms, account_ids=account_ids)
        
        return {
            "code": 200,
            "summary": summary,
            "videos": videos,
            "chartData": chart_data
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))



@router.get("/export", summary="导出分析数据")
async def export_analytics(
    startDate: Optional[str] = Query(None),
    endDate: Optional[str] = Query(None),
    platform: Optional[str] = Query(None, description="Legacy single platform"),
    platforms: Optional[List[str]] = Query(None, description="List of platforms"),
    accounts: Optional[List[str]] = Query(None, description="List of account IDs"),
    format: str = Query("csv", pattern="^(csv|excel)$", description="导出格式 csv 或 excel")
):
    """导出分析数据为 CSV（默认）或 Excel（若依赖存在）"""
    try:
        # Support legacy single platform param by adding it to list if present
        if platform and platform != "all":
            if not platforms:
                platforms = []
            platforms.append(platform)

        # Build account_ids list from accounts param
        account_ids = accounts

        videos = get_analytics_videos(
            DB_PATH,
            startDate,
            endDate,
            limit=10000,
            platforms=platforms,
            account_ids=account_ids
        )
        if format == "excel":
            try:
                import openpyxl  # type: ignore
                from openpyxl.styles import Font, Alignment  # type: ignore
            except ImportError:
                format = "csv"

        if format == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow([
                'ID', '视频ID', '标题', '平台', '视频链接',
                '发布日期', '播放量', '点赞量', '评论量', '收藏量', '最后更新'
            ])
            for video in videos:
                writer.writerow([
                    video.get('id'),
                    video.get('videoId'),
                    video.get('title'),
                    video.get('platform'),
                    video.get('videoUrl', ''),
                    video.get('publishDate'),
                    video.get('playCount', 0),
                    video.get('likeCount', 0),
                    video.get('commentCount', 0),
                    video.get('collectCount', 0),
                    video.get('lastUpdated')
                ])
            output.seek(0)
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="text/csv",
                headers={
                    "Content-Disposition": "attachment; filename=analytics_export.csv",
                    "Content-Type": "text/csv; charset=utf-8-sig"
                }
            )
        else:
            # Excel 导出
            import openpyxl  # type: ignore
            from openpyxl.styles import Font, Alignment  # type: ignore

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "视频数据"
            headers = [
                'ID', '视频ID', '标题', '平台', '视频链接',
                '发布日期', '播放量', '点赞量', '评论量', '收藏量', '最后更新'
            ]
            header_font = Font(bold=True)
            header_alignment = Alignment(horizontal='center', vertical='center')
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.alignment = header_alignment
            for row_idx, video in enumerate(videos, 2):
                ws.cell(row=row_idx, column=1, value=video.get('id'))
                ws.cell(row=row_idx, column=2, value=video.get('videoId'))
                ws.cell(row=row_idx, column=3, value=video.get('title'))
                ws.cell(row=row_idx, column=4, value=video.get('platform'))
                ws.cell(row=row_idx, column=5, value=video.get('videoUrl', ''))
                ws.cell(row=row_idx, column=6, value=video.get('publishDate'))
                ws.cell(row=row_idx, column=7, value=video.get('playCount', 0))
                ws.cell(row=row_idx, column=8, value=video.get('likeCount', 0))
                ws.cell(row=row_idx, column=9, value=video.get('commentCount', 0))
                ws.cell(row=row_idx, column=10, value=video.get('collectCount', 0))
                ws.cell(row=row_idx, column=11, value=video.get('lastUpdated'))
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            return StreamingResponse(
                iter([output.getvalue()]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=analytics_export.xlsx"}
            )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/video", summary="新增视频分析记录")
async def add_video_analytics(payload: dict):
    """新增视频分析记录"""
    try:
        video_id = insert_video_analytics(DB_PATH, payload)
        return {"code": 200, "msg": "Success", "videoId": video_id}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.put("/video/{video_id}", summary="更新视频分析数据")
async def update_video(video_id: int, payload: dict):
    """更新视频分析数据"""
    try:
        update_video_analytics(DB_PATH, video_id, payload)
        return {"code": 200, "msg": "Updated successfully"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/collect/{platform}", summary="Platform-specific analytics collection (Bilibili ready)")
async def collect_platform(platform: str, payload: Optional[CollectPayload] = Body(default=None)):
    platform = (platform or "").lower()
    account_ids = payload.account_ids if payload else None

    if platform == "bilibili":
        result = await _collect_bilibili_accounts(account_ids=account_ids)
        return {"success": True, "data": result, "message": "Bilibili collection completed"}
    if platform == "douyin":
        result = await _collect_douyin_tiktok_accounts(account_ids=account_ids)
        return {"success": True, "data": result, "message": "Douyin collection completed"}
    if platform == "xiaohongshu":
        result = await _collect_xhs_mediacrawler_accounts(account_ids=account_ids)
        return {"success": True, "data": result, "message": "Xiaohongshu collection completed"}
    if platform == "kuaishou":
        result = await _collect_kuaishou_mediacrawler_accounts(account_ids=account_ids)
        return {"success": True, "data": result, "message": "Kuaishou collection completed"}

    raise HTTPException(status_code=501, detail="Not implemented")


class DouyinTiktokCollectPayload(BaseModel):
    account_ids: Optional[List[str]] = Field(None, description="Douyin account IDs")
    sec_user_ids: Optional[Dict[str, str]] = Field(None, description="Map of account_id to sec_user_id")
    max_pages: int = Field(5, ge=1, le=50)
    page_size: int = Field(20, ge=1, le=50)


@router.post("/douyin-tiktok/collect", summary="Collect Douyin data via douyin_tiktok_api with per-account cookies")
async def collect_douyin_tiktok(payload: DouyinTiktokCollectPayload):
    result = await _collect_douyin_tiktok_accounts(
        account_ids=payload.account_ids,
        sec_user_ids=payload.sec_user_ids,
        max_pages=payload.max_pages,
        page_size=payload.page_size,
    )
    return {"success": True, "data": result, "message": "Douyin collection completed"}


class DouyinSecUidResolvePayload(BaseModel):
    account_ids: Optional[List[str]] = Field(None, description="Douyin account IDs")


@router.post("/douyin-tiktok/resolve-sec-uid", summary="Resolve Douyin sec_uid for accounts")
async def resolve_douyin_sec_uid(payload: DouyinSecUidResolvePayload):
    accounts = [
        acc for acc in cookie_manager.list_flat_accounts()
        if acc.get("platform") == "douyin"
    ]
    if payload.account_ids:
        accounts = [acc for acc in accounts if acc.get("account_id") in set(payload.account_ids)]

    resolved: Dict[str, Any] = {}
    failed: Dict[str, str] = {}

    for account in accounts:
        cookie_data = _read_cookie_file_for_account(account)
        sec_uid = _extract_sec_uid_from_cookie(cookie_data) or await _resolve_douyin_sec_uid(account, cookie_data)
        if sec_uid:
            _persist_douyin_sec_uid(account, sec_uid)
            resolved[str(account.get("account_id"))] = sec_uid
        else:
            failed[str(account.get("account_id"))] = "resolve_failed"

    return {"success": True, "data": {"resolved": resolved, "failed": failed}}


class DouyinIdToSecUidPayload(BaseModel):
    """抖音 ID 转 sec_uid 请求体"""
    user_ids: List[str] = Field(..., description="抖音号列表（数字ID）")
    cookie_header: Optional[str] = Field(None, description="可选的 Cookie 字符串")
    use_playwright: bool = Field(False, description="是否使用 Playwright 降级方案（较慢但更稳定）")


@router.post("/douyin/id-to-secuid", summary="抖音 ID 转 sec_uid", tags=["抖音工具"])
async def douyin_id_to_secuid(payload: DouyinIdToSecUidPayload):
    """
    通过抖音号（数字ID）批量查询 sec_uid
    
    支持三层降级策略：
    1. 搜索接口（最快，需要 a_bogus 签名）
    2. Playwright 模拟搜索（可选，较慢但稳定）
    3. 访问用户主页（降级方案）
    
    示例请求：
    ```json
    {
        "user_ids": ["12188823", "987654321"],
        "cookie_header": "可选的cookie字符串",
        "use_playwright": false
    }
    ```
    """
    resolved: Dict[str, str] = {}
    failed: Dict[str, str] = {}
    
    for user_id in payload.user_ids:
        try:
            sec_uid = await resolve_douyin_sec_uid(
                user_id=user_id,
                cookie_header=payload.cookie_header,
                use_playwright=payload.use_playwright
            )
            
            if sec_uid:
                resolved[user_id] = sec_uid
            else:
                failed[user_id] = "resolve_failed"
                
        except Exception as e:
            failed[user_id] = str(e)
    
    return {
        "success": len(failed) == 0,
        "data": {
            "resolved": resolved,
            "failed": failed,
            "total": len(payload.user_ids),
            "success_count": len(resolved),
            "failed_count": len(failed)
        }
    }


@router.get("/douyin/id-to-secuid/{user_id}", summary="单个抖音 ID 转 sec_uid", tags=["抖音工具"])
async def douyin_single_id_to_secuid(
    user_id: str,
    cookie: Optional[str] = Query(None, description="可选的 Cookie 字符串"),
    use_playwright: bool = Query(False, description="是否使用 Playwright")
):
    """
    通过抖音号（数字ID）查询单个 sec_uid
    
    示例：
    GET /api/v1/analytics/douyin/id-to-secuid/12188823
    """
    try:
        sec_uid = await resolve_douyin_sec_uid(
            user_id=user_id,
            cookie_header=cookie,
            use_playwright=use_playwright
        )
        
        if sec_uid:
            return {
                "success": True,
                "data": {
                    "user_id": user_id,
                    "sec_uid": sec_uid
                }
            }
        else:
            raise HTTPException(status_code=404, detail=f"无法解析用户 ID: {user_id}")
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


